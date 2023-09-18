import itertools
import os.path
import matplotlib.pyplot as plt
import pandas as pd
import numpy as np
from scipy.optimize import minimize
from scipy.signal import find_peaks
import openpyxl


class BassModel:
    def __init__(self):
        self.products_names = sheets_names
        self.number_products = len(self.products_names)
        self.t = np.empty(self.number_products, dtype=object)
        self.x_der = np.empty(self.number_products, dtype=object)
        self.x = np.empty(self.number_products, dtype=object)
        self.I = np.empty(self.number_products, dtype=object)
        self.M = np.empty(self.number_products, dtype=object)
        self.N_der = np.empty(self.number_products, dtype=object)

        # initiate initial parameters arrays
        self.initial_pm = np.zeros(self.number_products)
        self.initial_pi = np.zeros(self.number_products)
        self.initial_qi = np.zeros(self.number_products)
        self.initial_qm = np.zeros(self.number_products)
        self.initial_qim = np.zeros(self.number_products)
        self.initial_Ni = np.zeros(self.number_products)
        self.initial_Nm = np.zeros(self.number_products)
        self.minimized_params = None
        self.best_diff = np.inf

        # Initialize the estimation parameter values
        self.pm = np.zeros(self.number_products)
        self.pi = np.zeros(self.number_products)
        self.qi = np.zeros(self.number_products)
        self.qm = np.zeros(self.number_products)
        self.qim = np.zeros(self.number_products)
        self.Ni = np.zeros(self.number_products)
        self.Nm = np.zeros(self.number_products)

        # Initialize the estimation functions
        self.estimated_dI_dt = np.empty(self.number_products, dtype=object)
        self.estimated_dM_dt = np.empty(self.number_products, dtype=object)
        self.fit_Q_factor = np.zeros(self.number_products)
        self.fit_Q_factor_from_estimation_data = np.zeros(self.number_products)
        self.variance = np.zeros(self.number_products)

        # Initialize the saddle values
        self.had_a_saddle = [""] * self.number_products
        self.relative_depth = np.empty(self.number_products, dtype=int)
        self.saddle_duration = np.zeros(self.number_products)
        self.out_file = os.path.join(plots_path, "o.xlsx")
        self.out_columns = ["pm", "pi", "qi", "qm", "qim", "Ni", "Nm", "R-square"]
        self.run()

    def extract_data(self):
        self.initiate_outfile()
        for i, product in enumerate(self.products_names):
            # Read data from the Excel worksheet
            df = pd.read_excel(input_path, sheet_name=product)

            # real data
            df.columns = df.iloc[7]
            df.drop(df.index[:8], inplace=True)
            self.t[i] = np.array([int(x) for x in df["year"] if isinstance(x, int)
                                  and not pd.isna(x) and int(x) > 100])
            self.x_der[i] = np.array([float(x) for x in df["dx\dt"]
                                      if (isinstance(x, int) or isinstance(x, float)) and not pd.isna(x)])
            self.x[i] = np.array(
                [float(x) for x in df["x(t)"] if (isinstance(x, int) or isinstance(x, float)) and not pd.isna(x)])

            # arbitrary initial parameters values
            params = np.array(df["Parameters"])
            self.initial_pi[i] = float(params[2])
            self.initial_qi[i] = float(params[3])
            self.initial_qm[i] = float(params[4])
            self.initial_qim[i] = float(params[5])
            self.initial_Ni[i] = float(params[6])
            self.initial_Nm[i] = float(params[7])

            # estimated data
            df.columns = df.iloc[0]
            df.drop(df.index[0], inplace=True)
            self.I[i] = np.array([int(x) for x in df["I(t)"] if not pd.isna(x)])
            self.M[i] = np.array([int(x) for x in df["M(t)"] if not pd.isna(x)])
            self.N_der[i] = np.array([float(x) for x in df["d(I+M)/dt"] if not pd.isna(x)])
            self.calc_database_Q_factor(i)

    def run(self):
        self.extract_data()
        for i in range(self.number_products):
            self.estimate_parameters(i)
            self.get_saddle_and_dims(i)
        self.plot_saddle_dims()

    def find_single_dI_dt(self, I, pi, qi, Ni):
        return (pi + (qi * I / Ni)) * (Ni - I)

    def find_single_dM_dt(self, M, I, pm, qm, qim, Nm):
        return (pm + ((qm * M + qim * I) / Nm)) * (Nm - M)

    def gen_dI_dt(self, dx_dt, pi, qi, Ni):
        dI_dt = np.zeros(len(dx_dt))
        It = np.zeros(len(dx_dt))
        for j in range(len(dx_dt) - 1):
            dI_dt[j + 1] = self.find_single_dI_dt(It[j], pi, qi, Ni)
            It[j + 1] = It[j] + dI_dt[j + 1]
        return dI_dt, It

    def gen_dM_dt(self, arr, pm, qm, qim, Nm):
        dM_dt = np.zeros(len(arr))
        Mt = np.zeros(len(arr))
        for j in range(len(arr) - 1):
            dM_dt[j + 1] = self.find_single_dM_dt(Mt[j], arr[j], pm, qm, qim, Nm)
            Mt[j + 1] = dM_dt[j + 1] + Mt[j]
        return dM_dt, Mt

    def callback(self, params, i):
        pm, pi, qi, qm, qim, Ni, Nm = params
        dI_dt, It = self.gen_dI_dt(self.x_der[i], pi, qi, Ni)
        dM_dt, Mt = self.gen_dM_dt(It, pm, qm, qim, Nm)
        ssr = np.sum((self.x_der[i] - (dI_dt + dM_dt)) ** 2)
        sst = np.sum((self.x_der[i] - np.mean(self.x_der[i])) ** 2)
        r2 = 1 - (ssr / sst)
        if r2 > 1:
            return 10000
        return 1 - r2

    def estimate_parameters(self, i):
        initial = [self.initial_pm[i], self.initial_pi[i], self.initial_qi[i],
                   self.initial_qm[i], self.initial_qim[i], self.initial_Ni[i], self.initial_Nm[i]]
        bounds = [(0, 0.5), (0, 0.5), (0, 0.5), (0, 0.5), (0, 0.5), (0, np.inf), (0, np.inf)]
        result = minimize(self.callback, np.array(initial), bounds=bounds)
        self.pm[i], self.pi[i], self.qi[i], self.qm[i], self.qim[i], self.Ni[i], self.Nm[i] = result.x
        guesses = [[k * 0.1 + self.pm[i] for k in range(-2, 3, 1) if k * 0.1 + self.pm[i] <= 0.5],
                   [k * 0.1 + self.pi[i] for k in range(-2, 3, 1) if k * 0.1 + self.pi[i] <= 0.5],
                   [k * 0.1 + self.qi[i] for k in range(-2, 3, 1) if k * 0.1 + self.qi[i] <= 0.5],
                   [k * 0.1 + self.qm[i] for k in range(-2, 3, 1) if k * 0.1 + self.qm[i] <= 0.5],
                   [k * 0.1 + self.qim[i] for k in range(-2, 3, 1) if k * 0.1 + self.qim[i] <= 0.5],
                   [j + self.Ni[i] for j in range(-10000, 11000, 1000)],
                   [j + self.Nm[i] for j in range(-10000, 11000, 1000)]]
        self.run_layer(guesses, bounds, i)
        guesses[5] = [j + self.Ni[i] for j in range(-4000, 5000, 1000)]
        guesses[6] = [j + self.Nm[i] for j in range(-4000, 5000, 1000)]
        self.run_layer(guesses, bounds, i)
        guesses = [[k * 0.01 + self.pm[i] for k in range(-10, 10, 1) if k * 0.01 + self.pm[i] <= 0.5],
                   [k * 0.01 + self.pi[i] for k in range(-10, 10, 1) if k * 0.01 + self.pi[i] <= 0.5],
                   [k * 0.01 + self.qi[i] for k in range(-10, 10, 1) if k * 0.01 + self.qi[i] <= 0.5],
                   [k * 0.01 + self.qm[i] for k in range(-10, 10, 1) if k * 0.01 + self.qm[i] <= 0.5],
                   [k * 0.01 + self.qim[i] for k in range(-10, 10, 1) if k * 0.01 + self.qim[i] <= 0.5],
                   [j + self.Ni[i] for j in range(-1000, 1000, 100)],
                   [j + self.Nm[i] for j in range(-1000, 1000, 100)]]
        self.run_layer(guesses, bounds, i)
        guesses = [[k * 0.001 + self.pm[i] for k in range(-10, 10, 1) if k * 0.001 + self.pm[i] <= 0.5],
                   [k * 0.001 + self.pi[i] for k in range(-10, 10, 1) if k * 0.001 + self.pi[i] <= 0.5],
                   [k * 0.001 + self.qi[i] for k in range(-10, 10, 1) if k * 0.001 + self.qi[i] <= 0.5],
                   [k * 0.001 + self.qm[i] for k in range(-10, 10, 1) if k * 0.001 + self.qm[i] <= 0.5],
                   [k * 0.001 + self.qim[i] for k in range(-10, 10, 1) if k * 0.001 + self.qim[i] <= 0.5],
                   [j + self.Ni[i] for j in range(-100, 100, 10)],
                   [j + self.Nm[i] for j in range(-100, 100, 10)]]
        self.run_layer(guesses, bounds, i)

        # Set the predicted values of I and M
        self.estimated_dI_dt[i] = (self.pi[i] + (self.qi[i] * self.I[i] / self.Ni[i])) * (self.Ni[i] - self.I[i])
        self.estimated_dM_dt[i] = (((self.qm[i] * self.M[i] + self.qim[i] * self.I[i]) / self.Nm[i]) +
                                   self.pm[i]) * (self.Nm[i] - self.M[i])
        if not self.products_names[i] == "VideoGames":
            ssq = np.sum((self.x_der[i] - (self.estimated_dI_dt[i] + self.estimated_dM_dt[i])) ** 2)
        else:
            ssq = np.sum((self.x_der[i] - (self.estimated_dI_dt[i][4:] + self.estimated_dM_dt[i][4:])) ** 2)
        self.fit_Q_factor[i] = 1 - (ssq / self.variance[i])

    def initiate_outfile(self):
        wb = openpyxl.Workbook()
        for product in self.products_names:
            wb.create_sheet(product)
            wb[product].append(self.out_columns)
        wb.save(self.out_file)

    def run_layer(self, guesses, bounds, i):
        wb = openpyxl.load_workbook(self.out_file)
        for combination in itertools.product(*guesses):
            result = minimize(self.callback, combination, args=(i,), bounds=bounds)
            diff = result.fun
            if isinstance(diff, float) and 0 <= diff < self.best_diff:
                self.minimized_params = result.x
                self.best_diff = diff
                res = [par for par in result.x]
                res.append(1 - diff)
                wb[self.products_names[i]].append(res)
        self.pm[i] = self.minimized_params[0]
        self.pi[i] = self.minimized_params[1]
        self.qi[i] = self.minimized_params[2]
        self.qm[i] = self.minimized_params[3]
        self.qim[i] = self.minimized_params[4]
        self.Ni[i] = self.minimized_params[5]
        self.Nm[i] = self.minimized_params[6]
        wb.save(self.out_file)

    def get_saddle_and_dims(self, i):
        max_sales_points, _ = find_peaks(self.x_der[i])
        min_sales_points, _ = find_peaks(-self.x_der[i])
        if len(max_sales_points) == 0 and len(min_sales_points) == 0:
            self.had_a_saddle[i] = "No"
            self.plot_non_saddle_data(i)
            return
        d = 0
        saddle_start_idx = 0
        min_sales_saddle = 0
        for j in range((len(min_sales_points))):
            if 14 >= max_sales_points[j] < min_sales_points[j] \
                    and self.x_der[i][max_sales_points[j]] - self.x_der[i][min_sales_points[j]] > d:
                d = self.x_der[i][max_sales_points[j]] - self.x_der[i][min_sales_points[j]]
                saddle_start_idx = max_sales_points[j]
                min_sales_saddle = min_sales_points[j]
        if saddle_start_idx == 0:
            return
        self.had_a_saddle[i] = "Yes"
        self.relative_depth[i] = int((d / self.x_der[i][saddle_start_idx]) * 100)
        for j in range(saddle_start_idx + 1, len(self.x_der[i])):
            if self.x_der[i][j] >= self.x_der[i][saddle_start_idx]:
                self.saddle_duration[i] = int(self.t[i][j] - self.t[i][saddle_start_idx])
                break
        if self.relative_depth[i] < 20 or self.saddle_duration[i] <= 2:
            self.had_a_saddle[i] += "*"
        self.plot_saddle_data(i, saddle_start_idx, min_sales_saddle)

    def plot_non_saddle_data(self, i):
        plt.figure()
        plt.plot(self.t[i], self.x_der[i], label=r"real data - $\frac{dx}{dt}\ $")
        plt.xlabel("Time [Years]")
        plt.ylabel("Sales [A.U.]")
        plt.title(f"Sales Numbers of {self.products_names[i]} VS Time : Non-Saddle Case")
        plt.grid()
        plt.legend()
        # plt.show()
        plt.savefig(os.path.join(plots_path, f"dx_dt of {self.products_names[i]}.png"))
        plt.close()

    def plot_saddle_data(self, i, saddle_start_idx, min_sales_saddle):
        plt.figure()
        if self.products_names[i] == "VideoGames":
            t = np.zeros(len(self.x_der[i]))
            t[0] = self.t[i][0]
            for j in range(1, len(self.x_der[i])):
                t[j] = self.t[i][j + 4]
            plt.plot(t, self.x_der[i], label=r"real data - $\frac{dx}{dt}\ $")
            plt.plot(self.t[i][saddle_start_idx + 4], self.x_der[i][saddle_start_idx],
                     "X", color='g', label="saddle start peak")
            plt.plot(self.t[i][min_sales_saddle + 4], self.x_der[i][min_sales_saddle],
                     "X", color='r', label="saddle minima")
        else:
            plt.plot(self.t[i], self.x_der[i], label=r"real data - $\frac{dx}{dt}\ $")
            plt.plot(self.t[i][saddle_start_idx], self.x_der[i][saddle_start_idx],
                     "X", color='g', label="saddle start peak")
            plt.plot(self.t[i][min_sales_saddle], self.x_der[i][min_sales_saddle],
                     "X", color='r', label="saddle minima")
        plt.xlabel("Time [Years]")
        plt.ylabel("Sales [A.U.]")
        plt.title(f"Sales Numbers of {self.products_names[i]} VS Time : Saddle Case")
        plt.grid()
        plt.legend()
        # plt.show()
        plt.savefig(os.path.join(plots_path, f"dx_dt of {self.products_names[i]}.png"))
        plt.close()

    def calc_database_Q_factor(self, i):
        years = self.t[i][-1] - self.t[i][5] \
            if self.products_names[i] == "VideoGames" else self.t[i][-1] - self.t[i][0]
        df = years - 4
        ssq, var = 0, 0
        for j in range(1, len(self.x_der[i])):
            if self.products_names[i] == "VideoGames":
                ssq += (self.N_der[i][j + 4] - self.x_der[i][j]) ** 2
            else:
                ssq += (self.N_der[i][j] - self.x_der[i][j]) ** 2
            var += (self.x_der[i][j] - np.mean(self.x_der[i])) ** 2
        ssq /= df
        self.variance[i] = (var / years)
        self.fit_Q_factor_from_estimation_data[i] = 100 * (1 - (ssq / var))

    def calc_Q_factor(self, i, cur_dN_dt):
        years = self.t[i][-1] - self.t[i][5] \
            if self.products_names[i] == "VideoGames" else self.t[i][-1] - self.t[i][0]
        df = years - 4
        SOS_self_estimated_data = 0
        for j in range(1, len(self.x_der[i])):
            if self.products_names[i] == "VideoGames":
                SOS_self_estimated_data += (cur_dN_dt[j + 4] - self.x_der[i][j]) ** 2
            else:
                SOS_self_estimated_data += (cur_dN_dt[j] - self.x_der[i][j]) ** 2
        SOS_self_estimated_data /= df
        self.fit_Q_factor[i] = 100 * ((SOS_self_estimated_data / self.variance[i]) - 1)
        return self.fit_Q_factor[i]

    def plot_saddle_dims(self):
        table_data = []
        df_table = pd.DataFrame(columns=saddle_dims)
        for i in range(self.number_products):
            table_data.append([self.products_names[i], "{:.3f}".format(self.qi[i]), "{:.3f}".format(self.pi[i]),
                               "{:.3f}".format(self.pm[i]), "{:.3f}".format(self.qm[i]),
                               "{:.3f}".format(self.qim[i]),
                               int(self.Ni[i]), int(self.Nm[i]), self.had_a_saddle[i],
                               f"{self.relative_depth[i]}%" if self.had_a_saddle[i] != "No" else "---",
                               int(self.saddle_duration[i]) if self.had_a_saddle[i] != "No" else "---",
                               "{:.1f}".format(self.fit_Q_factor[i]) + "%",
                               "{:.1f}".format(self.fit_Q_factor_from_estimation_data[i]) + "%"])
            df_table.loc[i] = table_data[-1]
        table_data = np.array(table_data)
        df_table.to_csv(os.path.join(plots_path, "out.csv"), index=False)

        fig, ax = plt.subplots(figsize=(13, 2))
        table = ax.table(cellText=table_data, colLabels=saddle_dims, cellLoc='center', loc='center')
        header_cells = [table.get_celld()[0, j] for j in range(len(saddle_dims))]
        for cell in header_cells:
            cell.set_facecolor('lightblue')
            cell.get_text().set_weight('bold')
        for i in range(self.number_products):
            row_cells = [table.get_celld()[i + 1, j] for j in range(len(saddle_dims))]
            for cell in row_cells:
                if i % 2 == 0:
                    cell.set_facecolor('white')
                else:
                    cell.set_facecolor('#f2f2f2')
        for key, cell in table.get_celld().items():
            cell.set_linewidth(0)
        ax.axis('off')
        fig.suptitle("Estimated Parameters & Saddle Dimensions")
        table.auto_set_font_size(False)
        table.set_fontsize(6)
        plt.tight_layout()
        plt.savefig(os.path.join(plots_path, "parameters results.png"))
        plt.show()


def dI_dt_func(I, pi, qi, Ni):
    return (pi + qi * I / Ni) * (Ni - I)


def dM_dt_func(N, t, qm, Ni, Nm, qim):
    I, M = N
    (qm * M / (Ni + Nm) + qim * I / (Ni + Nm)) * (Nm - M)


def ode(N, pm, pi, qi, qm, qim, Ni, Nm):
    I, M = N
    return (pi + qi * I / Ni) * (Ni - I) + ((qm * (M / (Ni + Nm))) + (qim * (I / (Ni + Nm))) + pm) * (Nm - M)


course_path = ""

input_path = course_path + "Exercises/ChasmEconomist2003c - for estimation.xlsx"
sheets_names = ["PC's", "Mobile Phone", "VCR's", "VideoGames", "CD Players", "Answering mc", "Cordless Phone"]
saddle_dims = ["Product", r"Estimated $q_i$", r"Estimated $p_i$", r"Estimated $p_m$", r"Estimated $q_m$",
               r"Estimated $q_im$", r"Estimated $N_i$", r"Estimated $N_m$", "Had a saddle?", "Relative depth",
               "Saddle duration", "R-square self est.", "R-square est. data"]
plots_path = course_path + "plots/Ex3/"

if __name__ == "__main__":
    p = BassModel()
